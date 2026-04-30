"""
Vistas adicionales para Pagos, Historial de matriculados y Estudiantes.

Diseño:
- Todas usan el decorador @matricula_requerida (admin + asesor pueden ver).
- Las exportaciones a Excel usan openpyxl y devuelven un HttpResponse con el archivo.
- Filtros por GET querystring (q, curso, modalidad, estado, año, mes).
"""

from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST

from .forms import AbonoForm
from .models import Abono, Curso, Estudiante, Matricula
from .permisos import matricula_requerida


# ═════════════════════════════════════════════════════════════════
# Constantes
# ═════════════════════════════════════════════════════════════════

MESES_ES = [
    '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
]


# ═════════════════════════════════════════════════════════════════
# Helpers de Excel
# ═════════════════════════════════════════════════════════════════

def _build_excel_response(filename, sheet_name, headers, rows, totals=None):
    """
    Genera un .xlsx en memoria y lo devuelve como HttpResponse para descarga.

    headers: lista de strings (encabezados de columna)
    rows: lista de listas (datos)
    totals: dict opcional {col_idx_0based: total} para fila final
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel limita a 31 chars

    # ── Estilos ──
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1A237E')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )
    total_font = Font(bold=True, color='1A237E', size=11)
    total_fill = PatternFill('solid', fgColor='FFF8E1')

    # ── Título de la hoja en fila 1 ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=sheet_name)
    title_cell.font = Font(bold=True, size=14, color='1A237E')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    # ── Encabezados en fila 2 ──
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 30

    # ── Datos a partir de fila 3 ──
    for row_idx, row_data in enumerate(rows, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=False)

    # ── Fila de totales ──
    if totals:
        total_row_idx = len(rows) + 3
        # Etiqueta "TOTAL" en la primera columna
        cell = ws.cell(row=total_row_idx, column=1, value='TOTAL')
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right', vertical='center')

        for col_idx_0, value in totals.items():
            cell = ws.cell(row=total_row_idx, column=col_idx_0 + 1, value=value)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    # ── Auto-ancho aproximado por columna ──
    for col_idx, header in enumerate(headers, start=1):
        max_length = len(str(header))
        for row_data in rows:
            if col_idx - 1 < len(row_data):
                value = row_data[col_idx - 1]
                if value is not None:
                    max_length = max(max_length, len(str(value)))
        ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = min(max_length + 3, 40)

    # ── Congelar encabezados ──
    ws.freeze_panes = 'A3'

    # ── Devolver como HttpResponse ──
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _filtrar_matriculas(request):
    """
    Aplica filtros comunes a una queryset de Matricula según los GET params.
    Devuelve (queryset filtrada, dict de filtros aplicados).
    """
    qs = Matricula.objects.select_related(
        'estudiante', 'curso', 'curso__categoria', 'jornada', 'registrado_por'
    )

    estado = request.GET.get('estado', '').strip()
    curso_id = request.GET.get('curso', '').strip()
    modalidad = request.GET.get('modalidad', '').strip()
    anio = request.GET.get('anio', '').strip()
    mes = request.GET.get('mes', '').strip()
    q = request.GET.get('q', '').strip()

    if curso_id:
        qs = qs.filter(curso_id=curso_id)
    if modalidad in ('presencial', 'online'):
        qs = qs.filter(modalidad=modalidad)
    if anio.isdigit():
        qs = qs.filter(fecha_matricula__year=int(anio))
    if mes.isdigit() and 1 <= int(mes) <= 12:
        qs = qs.filter(fecha_matricula__month=int(mes))
    if q:
        qs = qs.filter(
            Q(estudiante__cedula__icontains=q)
            | Q(estudiante__apellidos__icontains=q)
            | Q(estudiante__nombres__icontains=q)
            | Q(curso__nombre__icontains=q)
        )

    # Filtro por estado (Pagado/Parcial/Pendiente) — se hace en Python
    # porque saldo es una @property, no un campo de DB.
    # Para mantener qs como queryset, traduzco el estado a condiciones:
    if estado == 'Pagado':
        qs = qs.filter(valor_pagado__gte=models_F('valor_curso'))
    elif estado == 'Parcial':
        qs = qs.filter(valor_pagado__gt=0, valor_pagado__lt=models_F('valor_curso'))
    elif estado == 'Pendiente':
        qs = qs.filter(Q(valor_pagado=0) | Q(valor_pagado__isnull=True))

    return qs, {
        'estado': estado,
        'curso': curso_id,
        'modalidad': modalidad,
        'anio': anio,
        'mes': mes,
        'q': q,
    }


# Importación tardía para evitar circular imports en algunos casos
from django.db.models import F as models_F


# ═════════════════════════════════════════════════════════════════
# Pagos
# ═════════════════════════════════════════════════════════════════

@matricula_requerida
def pagos_lista(request):
    """
    Vista centrada en lo financiero. Permite filtrar matrículas por:
    - Estado de pago: Pagado, Parcial, Pendiente
    - Curso (buscador)
    - Modalidad
    """
    qs, filtros = _filtrar_matriculas(request)
    qs = qs.order_by('-fecha_matricula', '-id')

    # Estadísticas globales (con los filtros aplicados, excepto el de estado)
    qs_sin_estado = Matricula.objects.select_related('curso').all()
    if filtros['curso']:
        qs_sin_estado = qs_sin_estado.filter(curso_id=filtros['curso'])
    if filtros['modalidad']:
        qs_sin_estado = qs_sin_estado.filter(modalidad=filtros['modalidad'])
    if filtros['anio'].isdigit():
        qs_sin_estado = qs_sin_estado.filter(fecha_matricula__year=int(filtros['anio']))

    totales = {
        'total_matriculas': qs_sin_estado.count(),
        'total_facturado': qs_sin_estado.aggregate(s=Sum('valor_curso'))['s'] or Decimal('0.00'),
        'total_cobrado': qs_sin_estado.aggregate(s=Sum('valor_pagado'))['s'] or Decimal('0.00'),
    }
    totales['total_pendiente'] = totales['total_facturado'] - totales['total_cobrado']

    # Conteo por estado
    todos_los_pagos = list(qs_sin_estado.values('valor_curso', 'valor_pagado'))
    conteo_estado = {'Pagado': 0, 'Parcial': 0, 'Pendiente': 0}
    for p in todos_los_pagos:
        vc = p['valor_curso'] or Decimal('0.00')
        vp = p['valor_pagado'] or Decimal('0.00')
        if vp >= vc and vc > 0:
            conteo_estado['Pagado'] += 1
        elif vp > 0:
            conteo_estado['Parcial'] += 1
        else:
            conteo_estado['Pendiente'] += 1

    cursos = Curso.objects.filter(activo=True).order_by('nombre')
    anios = sorted(
        set(Matricula.objects.dates('fecha_matricula', 'year').values_list('fecha_matricula__year', flat=True)),
        reverse=True
    )

    return render(request, 'pagos/lista.html', {
        'matriculas': qs,
        'cursos': cursos,
        'anios': anios,
        'filtros': filtros,
        'totales': totales,
        'conteo_estado': conteo_estado,
    })


@matricula_requerida
def pagos_export(request):
    """Descarga los pagos filtrados como Excel."""
    qs, filtros = _filtrar_matriculas(request)
    qs = qs.order_by('-fecha_matricula', '-id')

    headers = [
        'Fecha matrícula', 'Cédula', 'Apellidos', 'Nombres',
        'Curso', 'Categoría', 'Modalidad', 'Sede / Plataforma',
        'Valor curso', 'Valor pagado', 'Saldo', 'Estado',
    ]

    rows = []
    total_curso = Decimal('0.00')
    total_pagado = Decimal('0.00')
    total_saldo = Decimal('0.00')

    for m in qs:
        rows.append([
            m.fecha_matricula.strftime('%d/%m/%Y') if m.fecha_matricula else '',
            m.estudiante.cedula,
            m.estudiante.apellidos,
            m.estudiante.nombres,
            m.curso.nombre,
            m.curso.categoria.nombre if m.curso.categoria else '—',
            m.get_modalidad_display(),
            m.sede,
            float(m.valor_curso or 0),
            float(m.valor_pagado or 0),
            float(m.saldo or 0),
            m.estado_pago,
        ])
        total_curso += m.valor_curso or Decimal('0.00')
        total_pagado += m.valor_pagado or Decimal('0.00')
        total_saldo += m.saldo or Decimal('0.00')

    totals = {
        8: float(total_curso),
        9: float(total_pagado),
        10: float(total_saldo),
    }

    fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
    sufijo = ''
    if filtros['estado']:
        sufijo += f"_{filtros['estado'].lower()}"
    if filtros['anio']:
        sufijo += f"_{filtros['anio']}"
    filename = f'pagos{sufijo}_{fecha_str}.xlsx'

    return _build_excel_response(
        filename=filename,
        sheet_name='Reporte de Pagos',
        headers=headers,
        rows=rows,
        totals=totals,
    )


# ═════════════════════════════════════════════════════════════════
# Historial de matriculados (por año / mes)
# ═════════════════════════════════════════════════════════════════

@matricula_requerida
def historial_lista(request):
    """
    Historial de matrículas agrupado por año y mes.
    Permite filtrar por año, mes, curso y modalidad.
    """
    qs, filtros = _filtrar_matriculas(request)
    qs = qs.order_by('-fecha_matricula', '-id')

    # Agrupar por año → mes → matrículas
    grupos = defaultdict(lambda: defaultdict(list))
    totales_por_anio = defaultdict(lambda: {'count': 0, 'facturado': Decimal('0.00'), 'cobrado': Decimal('0.00')})
    totales_por_mes = defaultdict(lambda: {'count': 0, 'facturado': Decimal('0.00'), 'cobrado': Decimal('0.00')})

    for m in qs:
        anio = m.fecha_matricula.year
        mes = m.fecha_matricula.month
        grupos[anio][mes].append(m)

        totales_por_anio[anio]['count'] += 1
        totales_por_anio[anio]['facturado'] += m.valor_curso or Decimal('0.00')
        totales_por_anio[anio]['cobrado'] += m.valor_pagado or Decimal('0.00')

        key = (anio, mes)
        totales_por_mes[key]['count'] += 1
        totales_por_mes[key]['facturado'] += m.valor_curso or Decimal('0.00')
        totales_por_mes[key]['cobrado'] += m.valor_pagado or Decimal('0.00')

    # Convertir a lista ordenada para el template
    estructura = []
    for anio in sorted(grupos.keys(), reverse=True):
        meses_dict = grupos[anio]
        meses_lista = []
        for mes in sorted(meses_dict.keys(), reverse=True):
            meses_lista.append({
                'numero': mes,
                'nombre': MESES_ES[mes],
                'matriculas': meses_dict[mes],
                'totales': totales_por_mes[(anio, mes)],
            })
        estructura.append({
            'anio': anio,
            'meses': meses_lista,
            'totales': totales_por_anio[anio],
        })

    cursos = Curso.objects.filter(activo=True).order_by('nombre')
    anios_disponibles = sorted(
        set(Matricula.objects.dates('fecha_matricula', 'year').values_list('fecha_matricula__year', flat=True)),
        reverse=True
    )

    return render(request, 'historial/lista.html', {
        'estructura': estructura,
        'cursos': cursos,
        'anios': anios_disponibles,
        'meses_es': MESES_ES,
        'filtros': filtros,
        'total_general': qs.count(),
    })


@matricula_requerida
def historial_export(request):
    """
    Descarga del historial como Excel. El archivo tiene una hoja por año
    (o una sola si se filtró por año específico).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    qs, filtros = _filtrar_matriculas(request)
    qs = qs.order_by('-fecha_matricula', '-id')

    # Agrupar por año
    por_anio = defaultdict(list)
    for m in qs:
        por_anio[m.fecha_matricula.year].append(m)

    if not por_anio:
        # Excel vacío con mensaje
        return _build_excel_response(
            filename='historial_vacio.xlsx',
            sheet_name='Historial',
            headers=['Sin datos'],
            rows=[['No hay matrículas con los filtros aplicados.']],
        )

    # Construir el archivo manualmente con varias hojas
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1A237E')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    month_font = Font(bold=True, color='1A237E', size=12)
    month_fill = PatternFill('solid', fgColor='FFF8E1')
    total_font = Font(bold=True, color='2E7D32', size=10)
    total_fill = PatternFill('solid', fgColor='E8F5E9')
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    headers = [
        'Fecha matrícula', 'Cédula', 'Apellidos y Nombres',
        'Curso', 'Modalidad', 'Categoría', 'Sede',
        'Valor curso', 'Pagado', 'Saldo', 'Estado',
    ]

    for anio in sorted(por_anio.keys(), reverse=True):
        ws = wb.create_sheet(title=f'Año {anio}')

        # Título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title = ws.cell(row=1, column=1, value=f'Historial de matrículas — {anio}')
        title.font = Font(bold=True, size=14, color='1A237E')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 24

        # Encabezados
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        ws.row_dimensions[2].height = 30

        # Agrupar por mes dentro del año
        por_mes = defaultdict(list)
        for m in por_anio[anio]:
            por_mes[m.fecha_matricula.month].append(m)

        current_row = 3
        total_anio_facturado = Decimal('0.00')
        total_anio_pagado = Decimal('0.00')

        for mes in sorted(por_mes.keys(), reverse=True):
            # Fila separadora del mes
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=len(headers),
            )
            mes_cell = ws.cell(
                row=current_row, column=1,
                value=f'▸ {MESES_ES[mes]} {anio}  ({len(por_mes[mes])} matrícula(s))'
            )
            mes_cell.font = month_font
            mes_cell.fill = month_fill
            mes_cell.alignment = Alignment(horizontal='left', vertical='center')
            current_row += 1

            mes_facturado = Decimal('0.00')
            mes_pagado = Decimal('0.00')

            for m in por_mes[mes]:
                row_data = [
                    m.fecha_matricula.strftime('%d/%m/%Y'),
                    m.estudiante.cedula,
                    m.estudiante.nombre_completo,
                    m.curso.nombre,
                    m.get_modalidad_display(),
                    m.curso.categoria.nombre if m.curso.categoria else '—',
                    m.sede,
                    float(m.valor_curso or 0),
                    float(m.valor_pagado or 0),
                    float(m.saldo or 0),
                    m.estado_pago,
                ]
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')
                current_row += 1
                mes_facturado += m.valor_curso or Decimal('0.00')
                mes_pagado += m.valor_pagado or Decimal('0.00')

            # Subtotal del mes
            for col_idx in range(1, 8):
                cell = ws.cell(row=current_row, column=col_idx, value='')
                cell.fill = total_fill
                cell.border = thin_border
            ws.cell(row=current_row, column=7, value='Subtotal mes:').font = total_font
            ws.cell(row=current_row, column=7).alignment = Alignment(horizontal='right', vertical='center')

            for col_idx, val in [(8, float(mes_facturado)), (9, float(mes_pagado)),
                                 (10, float(mes_facturado - mes_pagado))]:
                c = ws.cell(row=current_row, column=col_idx, value=val)
                c.font = total_font
                c.fill = total_fill
                c.border = thin_border
                c.alignment = Alignment(vertical='center')
            ws.cell(row=current_row, column=11, value='').fill = total_fill
            current_row += 2  # espacio extra antes del próximo mes

            total_anio_facturado += mes_facturado
            total_anio_pagado += mes_pagado

        # Total del año
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_idx, value='')
            cell.fill = PatternFill('solid', fgColor='1A237E')
        ws.cell(row=current_row, column=7, value=f'TOTAL {anio}:').font = Font(bold=True, color='FFFFFF', size=11)
        ws.cell(row=current_row, column=7).fill = PatternFill('solid', fgColor='1A237E')
        ws.cell(row=current_row, column=7).alignment = Alignment(horizontal='right', vertical='center')

        for col_idx, val in [(8, float(total_anio_facturado)), (9, float(total_anio_pagado)),
                             (10, float(total_anio_facturado - total_anio_pagado))]:
            c = ws.cell(row=current_row, column=col_idx, value=val)
            c.font = Font(bold=True, color='FFFFFF', size=11)
            c.fill = PatternFill('solid', fgColor='1A237E')
            c.alignment = Alignment(vertical='center')

        # Auto-ancho
        for col_idx in range(1, len(headers) + 1):
            max_length = len(headers[col_idx - 1])
            for row_idx in range(3, current_row + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is not None:
                    max_length = max(max_length, len(str(v)))
            ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = min(max_length + 3, 38)

        ws.freeze_panes = 'A3'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f'historial_matriculados_{fecha_str}.xlsx'

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ═════════════════════════════════════════════════════════════════
# Estudiantes
# ═════════════════════════════════════════════════════════════════

@matricula_requerida
def estudiantes_lista(request):
    """
    Listado de estudiantes con búsqueda. Cada estudiante muestra el conteo
    de cursos matriculados y un enlace al detalle.
    """
    q = request.GET.get('q', '').strip()
    qs = Estudiante.objects.annotate(num_matriculas=Count('matriculas')).order_by('apellidos', 'nombres')

    if q:
        qs = qs.filter(
            Q(cedula__icontains=q)
            | Q(apellidos__icontains=q)
            | Q(nombres__icontains=q)
            | Q(correo__icontains=q)
            | Q(celular__icontains=q)
        )

    return render(request, 'estudiantes/lista.html', {
        'estudiantes': qs,
        'q': q,
        'total': qs.count(),
    })


@matricula_requerida
def estudiantes_por_curso(request):
    """
    Estudiantes agrupados por curso. Útil cuando se quiere ver la nómina
    completa de un curso específico.
    """
    curso_id = request.GET.get('curso', '').strip()
    modalidad = request.GET.get('modalidad', '').strip()

    cursos_qs = Curso.objects.filter(activo=True).order_by('nombre')

    grupos = []
    for curso in cursos_qs:
        if curso_id and str(curso.id) != curso_id:
            continue
        mat_qs = curso.matriculas.select_related('estudiante', 'jornada').order_by(
            'estudiante__apellidos', 'estudiante__nombres'
        )
        if modalidad in ('presencial', 'online'):
            mat_qs = mat_qs.filter(modalidad=modalidad)
        if mat_qs.exists() or not curso_id:
            grupos.append({
                'curso': curso,
                'matriculas': mat_qs,
                'total': mat_qs.count(),
            })

    # Ocultar cursos sin matriculados (excepto si se filtró por curso)
    if not curso_id:
        grupos = [g for g in grupos if g['total'] > 0]

    return render(request, 'estudiantes/por_curso.html', {
        'grupos': grupos,
        'cursos': cursos_qs,
        'curso_seleccionado': curso_id,
        'modalidad': modalidad,
    })


@matricula_requerida
def estudiante_detalle(request, pk):
    """Detalle de un estudiante con todas sus matrículas."""
    estudiante = get_object_or_404(Estudiante, pk=pk)
    matriculas = estudiante.matriculas.select_related(
        'curso', 'curso__categoria', 'jornada'
    ).order_by('-fecha_matricula')

    # Agrupar por año para el "historial"
    por_anio = defaultdict(list)
    for m in matriculas:
        por_anio[m.fecha_matricula.year].append(m)

    historial = []
    for anio in sorted(por_anio.keys(), reverse=True):
        items = por_anio[anio]
        historial.append({
            'anio': anio,
            'matriculas': items,
            'total_facturado': sum((m.valor_curso or Decimal('0.00')) for m in items),
            'total_pagado': sum((m.valor_pagado or Decimal('0.00')) for m in items),
        })

    return render(request, 'estudiantes/detalle.html', {
        'estudiante': estudiante,
        'matriculas': matriculas,
        'historial': historial,
        'total_matriculas': matriculas.count(),
    })


@matricula_requerida
def estudiantes_export(request):
    """
    Descarga el directorio de estudiantes como Excel.
    Si se pasa ?por_curso=1, genera una hoja por curso.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    por_curso = request.GET.get('por_curso', '') == '1'
    q = request.GET.get('q', '').strip()
    curso_id = request.GET.get('curso', '').strip()
    modalidad = request.GET.get('modalidad', '').strip()

    if por_curso:
        # Una hoja por curso (solo cursos con matriculados)
        wb = Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='1A237E')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD'),
        )

        headers = [
            'Cédula', 'Apellidos', 'Nombres', 'Edad',
            'Correo', 'Celular', 'Ciudad', 'Nivel',
            'Modalidad', 'Fecha matrícula', 'Valor', 'Pagado', 'Saldo', 'Estado',
        ]

        cursos_qs = Curso.objects.filter(activo=True).order_by('nombre')
        if curso_id and curso_id.isdigit():
            cursos_qs = cursos_qs.filter(id=int(curso_id))

        hojas_creadas = 0
        for curso in cursos_qs:
            mat_qs = curso.matriculas.select_related('estudiante').order_by(
                'estudiante__apellidos', 'estudiante__nombres'
            )
            if modalidad in ('presencial', 'online'):
                mat_qs = mat_qs.filter(modalidad=modalidad)

            if not mat_qs.exists():
                continue

            # Excel limita el nombre de hoja a 31 chars y prohíbe ciertos caracteres
            nombre_hoja = ''.join(c if c not in '\\/:*?[]' else '_' for c in curso.nombre)[:31]
            ws = wb.create_sheet(title=nombre_hoja)
            hojas_creadas += 1

            # Título
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            t = ws.cell(row=1, column=1, value=f'{curso.nombre} — {mat_qs.count()} estudiante(s)')
            t.font = Font(bold=True, size=14, color='1A237E')
            t.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 24

            # Encabezados
            for col_idx, h in enumerate(headers, start=1):
                c = ws.cell(row=2, column=col_idx, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = header_align
                c.border = thin_border
            ws.row_dimensions[2].height = 30

            for row_idx, m in enumerate(mat_qs, start=3):
                e = m.estudiante
                row_data = [
                    e.cedula, e.apellidos, e.nombres, e.edad or '',
                    e.correo or '', e.celular or '', e.ciudad or '',
                    e.get_nivel_formacion_display() if e.nivel_formacion else '',
                    m.get_modalidad_display(),
                    m.fecha_matricula.strftime('%d/%m/%Y') if m.fecha_matricula else '',
                    float(m.valor_curso or 0),
                    float(m.valor_pagado or 0),
                    float(m.saldo or 0),
                    m.estado_pago,
                ]
                for col_idx, val in enumerate(row_data, start=1):
                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.border = thin_border
                    c.alignment = Alignment(vertical='center')

            # Auto-ancho
            for col_idx in range(1, len(headers) + 1):
                max_length = len(headers[col_idx - 1])
                for row_idx in range(3, mat_qs.count() + 3):
                    v = ws.cell(row=row_idx, column=col_idx).value
                    if v is not None:
                        max_length = max(max_length, len(str(v)))
                ws.column_dimensions[
                    ws.cell(row=2, column=col_idx).column_letter
                ].width = min(max_length + 3, 35)

            ws.freeze_panes = 'A3'

        if hojas_creadas == 0:
            ws = wb.create_sheet(title='Sin datos')
            ws.cell(row=1, column=1, value='No hay estudiantes con los filtros aplicados.')

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f'estudiantes_por_curso_{fecha_str}.xlsx'
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # Modo plano: una sola hoja con todos los estudiantes
    estudiantes_qs = Estudiante.objects.annotate(
        num_matriculas=Count('matriculas')
    ).order_by('apellidos', 'nombres')

    if q:
        estudiantes_qs = estudiantes_qs.filter(
            Q(cedula__icontains=q)
            | Q(apellidos__icontains=q)
            | Q(nombres__icontains=q)
        )

    headers = [
        'Cédula', 'Apellidos', 'Nombres', 'Edad',
        'Correo', 'Celular', 'Ciudad', 'Nivel formación',
        'Título profesional', '# Matrículas', 'Cursos',
    ]

    rows = []
    for e in estudiantes_qs:
        cursos_str = ', '.join(
            sorted({m.curso.nombre for m in e.matriculas.all()})
        )
        rows.append([
            e.cedula, e.apellidos, e.nombres, e.edad or '',
            e.correo or '', e.celular or '', e.ciudad or '',
            e.get_nivel_formacion_display() if e.nivel_formacion else '',
            e.titulo_profesional or '',
            e.num_matriculas,
            cursos_str,
        ])

    fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f'estudiantes_{fecha_str}.xlsx'
    return _build_excel_response(
        filename=filename,
        sheet_name='Directorio de Estudiantes',
        headers=headers,
        rows=rows,
    )


@matricula_requerida
def estudiante_export(request, pk):
    """Descarga el historial individual de un estudiante."""
    estudiante = get_object_or_404(Estudiante, pk=pk)
    matriculas = estudiante.matriculas.select_related(
        'curso', 'curso__categoria', 'jornada'
    ).order_by('-fecha_matricula')

    headers = [
        'Año', 'Mes', 'Fecha matrícula', 'Curso', 'Modalidad',
        'Categoría', 'Sede', 'Valor', 'Pagado', 'Saldo', 'Estado',
    ]

    rows = []
    total_facturado = Decimal('0.00')
    total_pagado = Decimal('0.00')

    for m in matriculas:
        rows.append([
            m.fecha_matricula.year,
            MESES_ES[m.fecha_matricula.month],
            m.fecha_matricula.strftime('%d/%m/%Y'),
            m.curso.nombre,
            m.get_modalidad_display(),
            m.curso.categoria.nombre if m.curso.categoria else '—',
            m.sede,
            float(m.valor_curso or 0),
            float(m.valor_pagado or 0),
            float(m.saldo or 0),
            m.estado_pago,
        ])
        total_facturado += m.valor_curso or Decimal('0.00')
        total_pagado += m.valor_pagado or Decimal('0.00')

    totals = {
        7: float(total_facturado),
        8: float(total_pagado),
        9: float(total_facturado - total_pagado),
    }

    fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f'estudiante_{estudiante.cedula}_{fecha_str}.xlsx'
    return _build_excel_response(
        filename=filename,
        sheet_name=f'{estudiante.apellidos} {estudiante.nombres}'[:31],
        headers=headers,
        rows=rows,
        totals=totals,
    )


# ═════════════════════════════════════════════════════════════════
# Gestión de abonos por matrícula
# ═════════════════════════════════════════════════════════════════

@matricula_requerida
def matricula_abonos(request, pk):
    """
    Pantalla central de pagos de una matrícula:
    - Resumen (valor curso, pagado, saldo, estado)
    - Historial de abonos
    - Botón "Registrar abono" (modal)
    """
    matricula = get_object_or_404(
        Matricula.objects.select_related(
            'estudiante', 'curso', 'curso__categoria', 'jornada'
        ),
        pk=pk
    )
    abonos = matricula.abonos.select_related('registrado_por').order_by('-fecha', '-creado')

    # Saldo restante para el modal
    saldo_pendiente = matricula.saldo

    # Distribución por método (para mostrar resumen)
    dist_metodo = defaultdict(lambda: {'count': 0, 'total': Decimal('0.00')})
    for a in abonos:
        dist_metodo[a.get_metodo_display()]['count'] += 1
        dist_metodo[a.get_metodo_display()]['total'] += a.monto

    # Form pre-cargado para el modal (fecha=hoy, monto=saldo)
    form_inicial = AbonoForm(
        initial={
            'fecha': date.today(),
            'monto': saldo_pendiente if saldo_pendiente > 0 else None,
            'metodo': 'efectivo',
        },
        matricula=matricula,
    )

    return render(request, 'pagos/matricula_abonos.html', {
        'matricula': matricula,
        'abonos': abonos,
        'saldo_pendiente': saldo_pendiente,
        'dist_metodo': dict(dist_metodo),
        'form': form_inicial,
        'siguiente_recibo': Abono.generar_numero_recibo(),
    })


@matricula_requerida
@require_POST
def abono_crear(request, matricula_pk):
    """Crear un abono nuevo. Llamado desde el modal."""
    matricula = get_object_or_404(Matricula, pk=matricula_pk)
    form = AbonoForm(request.POST, matricula=matricula)

    if form.is_valid():
        abono = form.save(commit=False)
        abono.matricula = matricula
        abono.registrado_por = request.user
        abono.save()
        messages.success(
            request,
            f'Abono registrado: {abono.numero_recibo} por ${abono.monto}. '
            f'Nuevo saldo: ${matricula.saldo}.'
        )
    else:
        # Recopilar errores legibles (sin __all__ ni nombres internos)
        errores = []
        for field, errs in form.errors.items():
            prefijo = '' if field == '__all__' else f'{form.fields[field].label or field}: '
            for err in errs:
                errores.append(f'{prefijo}{err}')
        messages.error(
            request,
            'No se pudo registrar el abono. ' + ' / '.join(errores)
        )

    return redirect('academia:matricula_abonos', pk=matricula_pk)


@matricula_requerida
def abono_editar(request, matricula_pk, abono_pk):
    """Editar un abono existente."""
    matricula = get_object_or_404(Matricula, pk=matricula_pk)
    abono = get_object_or_404(Abono, pk=abono_pk, matricula=matricula)

    if request.method == 'POST':
        form = AbonoForm(request.POST, instance=abono, matricula=matricula)
        if form.is_valid():
            form.save()
            messages.success(request, f'Abono {abono.numero_recibo} actualizado.')
            return redirect('academia:matricula_abonos', pk=matricula_pk)
    else:
        form = AbonoForm(instance=abono, matricula=matricula)

    return render(request, 'pagos/abono_editar.html', {
        'form': form,
        'abono': abono,
        'matricula': matricula,
    })


@matricula_requerida
@require_POST
def abono_eliminar(request, matricula_pk, abono_pk):
    """Eliminar un abono y recalcular el total."""
    matricula = get_object_or_404(Matricula, pk=matricula_pk)
    abono = get_object_or_404(Abono, pk=abono_pk, matricula=matricula)
    numero = abono.numero_recibo
    monto = abono.monto
    abono.delete()
    messages.success(
        request,
        f'Abono {numero} eliminado (${monto}). Saldo recalculado: ${matricula.saldo}.'
    )
    return redirect('academia:matricula_abonos', pk=matricula_pk)


@matricula_requerida
def abonos_export(request):
    """
    Reporte de abonos en Excel — todos los abonos del periodo,
    con filtros por mes, año, método.
    """
    qs = Abono.objects.select_related(
        'matricula', 'matricula__estudiante', 'matricula__curso',
        'registrado_por',
    ).order_by('-fecha', '-creado')

    anio = request.GET.get('anio', '').strip()
    mes = request.GET.get('mes', '').strip()
    metodo = request.GET.get('metodo', '').strip()

    if anio.isdigit():
        qs = qs.filter(fecha__year=int(anio))
    if mes.isdigit() and 1 <= int(mes) <= 12:
        qs = qs.filter(fecha__month=int(mes))
    if metodo in ('efectivo', 'transferencia', 'tarjeta'):
        qs = qs.filter(metodo=metodo)

    headers = [
        'Nº Recibo', 'Fecha', 'Cédula', 'Estudiante', 'Curso',
        'Modalidad', 'Método', 'Monto', 'Valor curso', 'Saldo restante',
        'Registrado por', 'Observaciones',
    ]

    rows = []
    total_monto = Decimal('0.00')
    total_efectivo = Decimal('0.00')
    total_transf = Decimal('0.00')
    total_tarjeta = Decimal('0.00')

    for a in qs:
        m = a.matricula
        rows.append([
            a.numero_recibo,
            a.fecha.strftime('%d/%m/%Y'),
            m.estudiante.cedula,
            m.estudiante.nombre_completo,
            m.curso.nombre,
            m.get_modalidad_display(),
            a.get_metodo_display(),
            float(a.monto),
            float(m.valor_curso or 0),
            float(m.saldo or 0),
            (a.registrado_por.get_full_name() or a.registrado_por.username) if a.registrado_por else '—',
            a.observaciones or '',
        ])
        total_monto += a.monto
        if a.metodo == 'efectivo':
            total_efectivo += a.monto
        elif a.metodo == 'transferencia':
            total_transf += a.monto
        elif a.metodo == 'tarjeta':
            total_tarjeta += a.monto

    totals = {7: float(total_monto)}

    fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
    sufijo = ''
    if anio:
        sufijo += f'_{anio}'
    if mes:
        sufijo += f'_{mes:0>2}' if not mes.startswith('0') else f'_{mes}'
    if metodo:
        sufijo += f'_{metodo}'
    filename = f'abonos{sufijo}_{fecha_str}.xlsx'

    # Construir el archivo con totales por método al final
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte de Abonos'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1A237E')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )
    total_font = Font(bold=True, color='1A237E', size=11)
    total_fill = PatternFill('solid', fgColor='FFF8E1')
    method_font = Font(bold=True, color='2E7D32', size=11)
    method_fill = PatternFill('solid', fgColor='E8F5E9')

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    t = ws.cell(row=1, column=1, value='Reporte de Abonos')
    t.font = Font(bold=True, size=14, color='1A237E')
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    # Encabezados
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin
    ws.row_dimensions[2].height = 30

    # Datos
    for row_idx, row_data in enumerate(rows, start=3):
        for col_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = thin
            c.alignment = Alignment(vertical='center')

    # Total general
    total_row = len(rows) + 3
    ws.cell(row=total_row, column=6, value='TOTAL GENERAL:').font = total_font
    ws.cell(row=total_row, column=6).alignment = Alignment(horizontal='right')
    ws.cell(row=total_row, column=6).fill = total_fill
    ws.cell(row=total_row, column=6).border = thin
    c = ws.cell(row=total_row, column=8, value=float(total_monto))
    c.font = total_font
    c.fill = total_fill
    c.border = thin

    # Desglose por método
    metodo_row = total_row + 2
    ws.cell(row=metodo_row, column=1, value='💵 Por método de pago:').font = method_font
    metodo_row += 1
    for label, total in [
        ('Efectivo', total_efectivo),
        ('Transferencia', total_transf),
        ('Tarjeta', total_tarjeta),
    ]:
        ws.cell(row=metodo_row, column=1, value=label).font = method_font
        ws.cell(row=metodo_row, column=1).fill = method_fill
        ws.cell(row=metodo_row, column=1).border = thin
        c = ws.cell(row=metodo_row, column=2, value=float(total))
        c.font = method_font
        c.fill = method_fill
        c.border = thin
        metodo_row += 1

    # Auto-ancho
    for col_idx in range(1, len(headers) + 1):
        max_length = len(headers[col_idx - 1])
        for row_idx in range(3, len(rows) + 3):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                max_length = max(max_length, len(str(v)))
        ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = min(max_length + 3, 38)

    ws.freeze_panes = 'A3'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@matricula_requerida
def abono_recibo(request, abono_pk):
    """
    Vista del recibo individual (HTML imprimible).
    Cada abono tiene su comprobante.
    """
    abono = get_object_or_404(
        Abono.objects.select_related(
            'matricula', 'matricula__estudiante', 'matricula__curso',
            'registrado_por',
        ),
        pk=abono_pk
    )
    return render(request, 'pagos/recibo.html', {
        'abono': abono,
        'matricula': abono.matricula,
    })
